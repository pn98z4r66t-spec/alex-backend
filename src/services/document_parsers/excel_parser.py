"""
Excel Spreadsheet Parser
Extracts data, formulas, and metadata from Excel files (.xlsx, .xls)
"""
from openpyxl import load_workbook
from typing import Dict, List, Any
import logging

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parser for Excel spreadsheets (.xlsx)"""
    
    @staticmethod
    def parse(file_path: str, max_rows_per_sheet: int = 1000) -> Dict[str, Any]:
        """
        Parse Excel file and extract content
        
        Args:
            file_path: Path to .xlsx file
            max_rows_per_sheet: Maximum rows to extract per sheet (prevents memory issues)
            
        Returns:
            Dictionary containing:
                - text: Formatted text representation
                - sheets: List of sheet data
                - metadata: Workbook properties
                - has_formulas: Whether workbook contains formulas
        """
        try:
            result = {
                'text': '',
                'sheets': [],
                'metadata': {},
                'has_formulas': False,
                'success': True,
                'error': None
            }
            
            # Load workbook
            wb = load_workbook(file_path, data_only=False)
            
            # Extract metadata
            result['metadata'] = {
                'sheet_count': len(wb.sheetnames),
                'sheet_names': wb.sheetnames,
                'active_sheet': wb.active.title if wb.active else ''
            }
            
            # Extract properties if available
            if hasattr(wb, 'properties'):
                props = wb.properties
                result['metadata'].update({
                    'title': props.title or '',
                    'creator': props.creator or '',
                    'created': str(props.created) if props.created else '',
                    'modified': str(props.modified) if props.modified else '',
                })
            
            # Process each sheet
            all_text = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_data = {
                    'name': sheet_name,
                    'rows': [],
                    'row_count': sheet.max_row,
                    'column_count': sheet.max_column,
                    'has_formulas': False,
                    'cell_count': 0
                }
                
                # Extract data from rows (limit to prevent memory issues)
                rows_to_process = min(sheet.max_row, max_rows_per_sheet)
                
                for row_idx, row in enumerate(sheet.iter_rows(max_row=rows_to_process), 1):
                    row_data = []
                    for cell in row:
                        cell_value = cell.value
                        
                        # Check for formulas
                        if isinstance(cell_value, str) and cell_value.startswith('='):
                            sheet_data['has_formulas'] = True
                            result['has_formulas'] = True
                        
                        # Convert to string
                        row_data.append(str(cell_value) if cell_value is not None else '')
                    
                    # Only add non-empty rows
                    if any(cell.strip() for cell in row_data):
                        sheet_data['rows'].append(row_data)
                        sheet_data['cell_count'] += len(row_data)
                
                result['sheets'].append(sheet_data)
                
                # Add sheet content to text
                if sheet_data['rows']:
                    sheet_text = f"\n[Sheet: {sheet_name}]\n"
                    for row in sheet_data['rows'][:50]:  # Limit text output
                        sheet_text += ' | '.join(row) + '\n'
                    
                    if len(sheet_data['rows']) > 50:
                        sheet_text += f"... ({len(sheet_data['rows']) - 50} more rows)\n"
                    
                    all_text.append(sheet_text)
            
            # Combine all text
            result['text'] = '\n'.join(all_text)
            
            # Add statistics
            total_rows = sum(len(sheet['rows']) for sheet in result['sheets'])
            total_cells = sum(sheet['cell_count'] for sheet in result['sheets'])
            
            result['statistics'] = {
                'total_sheets': len(result['sheets']),
                'total_rows': total_rows,
                'total_cells': total_cells,
                'sheets_with_formulas': sum(1 for s in result['sheets'] if s['has_formulas'])
            }
            
            return result
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}")
            return {
                'success': False,
                'error': str(e),
                'text': '',
                'sheets': [],
                'metadata': {}
            }
    
    @staticmethod
    def get_summary(parsed_data: Dict[str, Any]) -> str:
        """
        Generate a human-readable summary of the Excel file
        
        Args:
            parsed_data: Output from parse() method
            
        Returns:
            Summary string
        """
        if not parsed_data.get('success'):
            return f"Failed to parse Excel file: {parsed_data.get('error')}"
        
        summary_parts = []
        
        # Basic info
        stats = parsed_data.get('statistics', {})
        metadata = parsed_data.get('metadata', {})
        
        summary_parts.append(
            f"Excel Workbook with {stats.get('total_sheets', 0)} sheet(s)"
        )
        
        # Sheet names
        if metadata.get('sheet_names'):
            summary_parts.append(f"Sheets: {', '.join(metadata['sheet_names'])}")
        
        # Content statistics
        summary_parts.append(
            f"Content: {stats.get('total_rows', 0)} rows, "
            f"{stats.get('total_cells', 0)} cells"
        )
        
        # Formulas
        if parsed_data.get('has_formulas'):
            summary_parts.append(
                f"Contains formulas in {stats.get('sheets_with_formulas', 0)} sheet(s)"
            )
        
        # Metadata
        if metadata.get('title'):
            summary_parts.append(f"Title: {metadata['title']}")
        if metadata.get('creator'):
            summary_parts.append(f"Creator: {metadata['creator']}")
        
        return '\n'.join(summary_parts)

